from __future__ import annotations

from copy import copy
from datetime import datetime
from io import BytesIO
from pathlib import Path
import re
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.worksheet.worksheet import Worksheet


BASE_DIR = Path(__file__).resolve().parent.parent
TEMPLATE_REPORT_XLSX = BASE_DIR / "templates" / "qf_pickup_report_template.xlsx"
REPORT_TEMPLATE_SHEET_NAME = "\ub9c8\uac10\uc790\ub8cc(\uc791\uc131)"

COL_TOTAL_QTY = "\ucd1d\uc218\ub7c9"
COL_CANCEL_COUNT = "\ucde8\uc18c\uac74\uc218"
COL_CANCEL_QTY = "\ucde8\uc18c\uc218\ub7c9"
COL_SETTLEMENT_EXCLUDED = "\uc815\uc0b0\uc81c\uc678"
COL_CANCEL_REASON = "\ucde8\uc18c\uc0ac\uc720"

TEXT_SUMMARY = "\ud569\uacc4"
TEXT_NOTE = "\ud2b9\uc774\uc0ac\ud56d"
TEXT_HEADER_PICKUP_CENTER = "\uc9d1\ud558\uc13c\ud130"
TEXT_SCHEDULED_QTY = "\uc608\uc815\uc218\ub7c9"
TEXT_PICKUP_QTY = "\ud53d\uc5c5\uc218\ub7c9"
TEXT_DIFF = "\ucc28\uc774"
TEXT_NO_DETAILS = "\uac10\uc18c\ub0b4\uc5ed \uc5c6\uc74c"
TEXT_SETTLEMENT_EXCLUDED = "\uc815\uc0b0 \uc81c\uc678"

CELL_MAP = {
    "report_date": "C4",
}

ROW_LAYOUT = {
    "header_row_marker": TEXT_HEADER_PICKUP_CENTER,
    "summary_marker": TEXT_SUMMARY,
    "note_marker": TEXT_NOTE,
    "detail_start_offset": 1,
    "detail_columns": {
        "route": "C",
        "milkrun_no": "D",
        "company_name": "E",
        "origin_center": "F",
        "decrease_display": "G",
    },
    "header_merge_columns": ("G", "H"),
    "detail_merge_columns": ("G", "H"),
    "summary_columns": {
        "label": "B",
        "scheduled_label": "C",
        "scheduled_qty": "D",
        "pickup_label": "E",
        "pickup_qty": "F",
        "diff_label": "G",
        "diff_qty": "H",
    },
    "note_columns": {
        "label": "A",
        "value": "C",
    },
    "note_default_height": 3,
    "sheet_max_column": 8,
}


def _safe_int(value: Any) -> int:
    try:
        if pd.isna(value):
            return 0
        return int(float(value))
    except Exception:
        return 0


def _safe_text(value: Any, default: str = "") -> str:
    if value is None:
        return default
    text = str(value).strip()
    return text if text else default


def _safe_numeric_series(df: pd.DataFrame, column: str) -> pd.Series:
    if column not in df.columns:
        return pd.Series([0] * len(df), index=df.index, dtype="float64")
    return pd.to_numeric(df[column], errors="coerce").fillna(0)


def _safe_text_series(df: pd.DataFrame, column: str) -> pd.Series:
    if column not in df.columns:
        return pd.Series([""] * len(df), index=df.index, dtype="object")
    return df[column].fillna("").astype(str).str.strip()


def _safe_bool_series(df: pd.DataFrame, column: str) -> pd.Series:
    if column not in df.columns:
        return pd.Series([False] * len(df), index=df.index, dtype="bool")
    return df[column].apply(_safe_bool_value)


def _safe_bool_value(value: Any) -> bool:
    try:
        if pd.isna(value):
            return False
    except Exception:
        pass

    if isinstance(value, str):
        normalized = value.strip().lower()
        if normalized in {"", "0", "false", "n", "no", "off"}:
            return False
        if normalized in {"1", "true", "y", "yes", "on"}:
            return True
    return bool(value)


def _normalize_marker_text(value: Any) -> str:
    return _safe_text(value).replace(" ", "")


def _parse_report_date(base_date_str: str) -> datetime | None:
    text = _safe_text(base_date_str)
    if not text:
        return None

    try:
        return pd.to_datetime(text, errors="raise").to_pydatetime()
    except Exception:
        match = re.search(r"(20\d{2})[-_./]?(0[1-9]|1[0-2])[-_./]?(0[1-9]|[12]\d|3[01])", text)
        if not match:
            return None
        yyyy, mm, dd = match.groups()
        return datetime(int(yyyy), int(mm), int(dd))


def _build_note_subject(row: dict[str, Any]) -> str:
    route_prefix = _safe_text(row.get("route_prefix"))
    house_order = _safe_int(row.get("house_order"))
    route = _safe_text(row.get("route"))
    company_name = _safe_text(row.get("company_name"))
    milkrun_no = _safe_text(row.get("milkrun_no"))

    route_label = ""
    if route_prefix or house_order > 0:
        route_label = f"{route_prefix}{house_order if house_order > 0 else ''}".strip()
    elif route:
        route_label = route
    elif milkrun_no:
        route_label = milkrun_no

    if route_label and company_name:
        return f"{route_label} {company_name}"
    return company_name or route_label or milkrun_no


def build_report_export_df(cancel_df: pd.DataFrame) -> pd.DataFrame:
    if len(cancel_df) == 0:
        return cancel_df.copy()

    work_df = cancel_df.copy()
    mask = (
        (_safe_numeric_series(work_df, COL_CANCEL_COUNT) > 0)
        | (_safe_numeric_series(work_df, COL_CANCEL_QTY) > 0)
        | _safe_bool_series(work_df, COL_SETTLEMENT_EXCLUDED)
        | _safe_text_series(work_df, COL_CANCEL_REASON).ne("")
    )
    return work_df[mask].copy().reset_index(drop=True)


def build_report_export_payload(
    grouped_delivery: pd.DataFrame,
    report_export_df: pd.DataFrame,
    base_date_str: str,
) -> dict[str, Any]:
    grouped_df = grouped_delivery.copy() if isinstance(grouped_delivery, pd.DataFrame) else pd.DataFrame()
    export_df = report_export_df.copy() if isinstance(report_export_df, pd.DataFrame) else pd.DataFrame()

    planned_qty = _safe_int(
        (
            _safe_numeric_series(grouped_df, "ae_sum")
            + _safe_numeric_series(grouped_df, "af_sum")
            + _safe_numeric_series(grouped_df, "ag_sum")
        ).sum()
    )
    cancel_qty_total = _safe_int(_safe_numeric_series(export_df, COL_CANCEL_QTY).sum())
    pickup_qty = planned_qty - cancel_qty_total
    difference_qty = pickup_qty - planned_qty

    detail_rows: list[dict[str, Any]] = []
    note_lines: list[str] = []

    for _, row in export_df.iterrows():
        total_qty = _safe_int(row.get(COL_TOTAL_QTY, 0))
        cancel_count = _safe_int(row.get(COL_CANCEL_COUNT, 0))
        cancel_qty = _safe_int(row.get(COL_CANCEL_QTY, 0))
        settlement_excluded = _safe_bool_value(row.get(COL_SETTLEMENT_EXCLUDED, False))
        cancel_reason = _safe_text(row.get(COL_CANCEL_REASON))
        adjusted_qty = max(0, total_qty - cancel_qty)

        if cancel_qty > 0:
            decrease_display = f"{total_qty} > {adjusted_qty}"
        elif cancel_count > 0:
            decrease_display = f"\ucde8\uc18c {cancel_count}\uac74"
        elif settlement_excluded:
            decrease_display = TEXT_SETTLEMENT_EXCLUDED
        else:
            decrease_display = ""

        row_payload = {
            "route": _safe_text(row.get("route")),
            "milkrun_no": _safe_text(row.get("milkrun_no")),
            "company_name": _safe_text(row.get("company_name")),
            "origin_center": _safe_text(row.get("origin_center")),
            "decrease_display": decrease_display,
            "route_prefix": _safe_text(row.get("route_prefix")),
            "house_order": row.get("house_order"),
        }
        detail_rows.append(row_payload)

        note_parts: list[str] = []
        if settlement_excluded:
            note_parts.append(TEXT_SETTLEMENT_EXCLUDED)
        if cancel_reason:
            note_parts.append(cancel_reason)
        if note_parts:
            subject = _build_note_subject({**row_payload, **row.to_dict()})
            if subject:
                note_lines.append(f"{subject}: {' / '.join(note_parts)}")
            else:
                note_lines.append(" / ".join(note_parts))

    note_text = "\n".join(note_lines) if note_lines else TEXT_NO_DETAILS

    return {
        "report_date": _parse_report_date(base_date_str),
        "report_date_text": _safe_text(base_date_str),
        "planned_qty": planned_qty,
        "pickup_qty": pickup_qty,
        "difference_qty": difference_qty,
        "detail_rows": detail_rows,
        "note_text": note_text,
    }


def _find_cell_by_marker(ws: Worksheet, marker_text: str):
    normalized_marker = _normalize_marker_text(marker_text)
    for row in ws.iter_rows():
        for cell in row:
            if _normalize_marker_text(cell.value) == normalized_marker:
                return cell
    return None


def _find_note_block_height(ws: Worksheet, note_row: int) -> int:
    for merged_range in ws.merged_cells.ranges:
        if merged_range.min_row <= note_row <= merged_range.max_row and merged_range.min_col == 1:
            return merged_range.max_row - merged_range.min_row + 1
    return int(ROW_LAYOUT["note_default_height"])


def _copy_row_style(ws: Worksheet, source_row: int, target_row: int) -> None:
    source_dimension = ws.row_dimensions[source_row]
    target_dimension = ws.row_dimensions[target_row]
    target_dimension.height = source_dimension.height
    target_dimension.hidden = source_dimension.hidden
    target_dimension.outlineLevel = source_dimension.outlineLevel
    target_dimension.ht = source_dimension.ht

    for col_index in range(1, int(ROW_LAYOUT["sheet_max_column"]) + 1):
        source_cell = ws.cell(source_row, col_index)
        target_cell = ws.cell(target_row, col_index)

        if source_cell.has_style:
            target_cell._style = copy(source_cell._style)
            target_cell.font = copy(source_cell.font)
            target_cell.fill = copy(source_cell.fill)
            target_cell.border = copy(source_cell.border)
            target_cell.alignment = copy(source_cell.alignment)
            target_cell.protection = copy(source_cell.protection)
            target_cell.number_format = source_cell.number_format
        if source_cell.comment is not None:
            target_cell.comment = copy(source_cell.comment)
        target_cell.value = None


def _clear_range(ws: Worksheet, start_row: int, end_row: int, start_col: int, end_col: int) -> None:
    if end_row < start_row:
        return
    for row_index in range(start_row, end_row + 1):
        for col_index in range(start_col, end_col + 1):
            cell = ws.cell(row_index, col_index)
            if isinstance(cell, MergedCell):
                continue
            cell.value = None


def _unmerge_dynamic_ranges(ws: Worksheet, start_row: int, end_row: int) -> None:
    for merged_range in list(ws.merged_cells.ranges):
        if merged_range.max_row < start_row:
            continue
        if merged_range.max_col > int(ROW_LAYOUT["sheet_max_column"]):
            continue
        ws.unmerge_cells(str(merged_range))


def fill_report_template(workbook, payload: dict[str, Any]):
    if REPORT_TEMPLATE_SHEET_NAME not in workbook.sheetnames:
        raise ValueError(f"missing sheet: {REPORT_TEMPLATE_SHEET_NAME}")

    ws = workbook[REPORT_TEMPLATE_SHEET_NAME]

    header_cell = _find_cell_by_marker(ws, str(ROW_LAYOUT["header_row_marker"]))
    summary_cell = _find_cell_by_marker(ws, str(ROW_LAYOUT["summary_marker"]))
    note_cell = _find_cell_by_marker(ws, str(ROW_LAYOUT["note_marker"]))
    if header_cell is None or summary_cell is None or note_cell is None:
        raise ValueError("report template markers are missing")

    header_row = header_cell.row
    detail_start_row = header_row + int(ROW_LAYOUT["detail_start_offset"])
    summary_row = summary_cell.row
    note_row = note_cell.row
    note_height = _find_note_block_height(ws, note_row)

    template_detail_row_count = summary_row - detail_start_row
    if template_detail_row_count < 1:
        raise ValueError("report template does not have a writable detail area")

    detail_rows = list(payload.get("detail_rows", []) or [])
    extra_rows = max(0, len(detail_rows) - template_detail_row_count)
    detail_style_row = max(detail_start_row, summary_row - 1)

    if extra_rows > 0:
        ws.insert_rows(summary_row, amount=extra_rows)
        for row_index in range(summary_row, summary_row + extra_rows):
            _copy_row_style(ws, detail_style_row, row_index)
        summary_row += extra_rows
        note_row += extra_rows

    note_end_row = note_row + note_height - 1

    section_label = ws.cell(header_row, 1).value
    pickup_center_value = ws.cell(detail_start_row, 2).value
    header_decrease_label = ws.cell(header_row, 7).value
    note_label = ws.cell(note_row, 1).value

    _unmerge_dynamic_ranges(ws, header_row, note_end_row)

    _clear_range(ws, detail_start_row, summary_row - 1, 2, int(ROW_LAYOUT["sheet_max_column"]))
    _clear_range(ws, summary_row, summary_row, 2, int(ROW_LAYOUT["sheet_max_column"]))
    _clear_range(ws, note_row, note_end_row, 1, int(ROW_LAYOUT["sheet_max_column"]))

    report_date_value = payload.get("report_date") or payload.get("report_date_text", "")
    ws[CELL_MAP["report_date"]] = report_date_value

    ws.cell(header_row, 1).value = section_label
    if summary_row > detail_start_row:
        ws.cell(detail_start_row, 2).value = pickup_center_value
    ws.cell(header_row, 7).value = header_decrease_label

    header_merge_start, header_merge_end = ROW_LAYOUT["header_merge_columns"]
    ws.merge_cells(f"{header_merge_start}{header_row}:{header_merge_end}{header_row}")
    ws.merge_cells(f"A{header_row}:A{summary_row}")
    if summary_row - 1 > detail_start_row:
        ws.merge_cells(f"B{detail_start_row}:B{summary_row - 1}")

    detail_merge_start, detail_merge_end = ROW_LAYOUT["detail_merge_columns"]
    for row_index in range(detail_start_row, summary_row):
        ws.merge_cells(f"{detail_merge_start}{row_index}:{detail_merge_end}{row_index}")

    summary_columns = dict(ROW_LAYOUT["summary_columns"])
    ws[f"{summary_columns['label']}{summary_row}"] = TEXT_SUMMARY
    ws[f"{summary_columns['scheduled_label']}{summary_row}"] = TEXT_SCHEDULED_QTY
    ws[f"{summary_columns['scheduled_qty']}{summary_row}"] = _safe_int(payload.get("planned_qty", 0))
    ws[f"{summary_columns['pickup_label']}{summary_row}"] = TEXT_PICKUP_QTY
    ws[f"{summary_columns['pickup_qty']}{summary_row}"] = _safe_int(payload.get("pickup_qty", 0))
    ws[f"{summary_columns['diff_label']}{summary_row}"] = TEXT_DIFF
    ws[f"{summary_columns['diff_qty']}{summary_row}"] = _safe_int(payload.get("difference_qty", 0))

    detail_columns = dict(ROW_LAYOUT["detail_columns"])
    for offset, row_data in enumerate(detail_rows):
        row_index = detail_start_row + offset
        ws[f"{detail_columns['route']}{row_index}"] = row_data.get("route", "")
        ws[f"{detail_columns['milkrun_no']}{row_index}"] = row_data.get("milkrun_no", "")
        ws[f"{detail_columns['company_name']}{row_index}"] = row_data.get("company_name", "")
        ws[f"{detail_columns['origin_center']}{row_index}"] = row_data.get("origin_center", "")
        ws[f"{detail_columns['decrease_display']}{row_index}"] = row_data.get("decrease_display", "")

    note_columns = dict(ROW_LAYOUT["note_columns"])
    ws[f"{note_columns['label']}{note_row}"] = note_label or TEXT_NOTE
    ws[f"{note_columns['value']}{note_row}"] = _safe_text(payload.get("note_text"), TEXT_NO_DETAILS)

    ws.merge_cells(f"A{note_row}:B{note_end_row}")
    ws.merge_cells(f"C{note_row}:H{note_end_row}")

    note_value_cell = ws[f"{note_columns['value']}{note_row}"]
    note_alignment = copy(note_value_cell.alignment)
    note_alignment.wrap_text = True
    note_value_cell.alignment = note_alignment

    return workbook


def make_report_xlsx_bytes(payload: dict[str, Any], template_path: Path | None = None) -> bytes:
    effective_template_path = Path(template_path) if template_path is not None else TEMPLATE_REPORT_XLSX
    if not effective_template_path.exists():
        raise FileNotFoundError("\ud15c\ud50c\ub9bf \ud30c\uc77c\uc774 \ud544\uc694\ud569\ub2c8\ub2e4")

    workbook = load_workbook(effective_template_path)
    fill_report_template(workbook, payload)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def build_report_export_filename(base_date_str: str) -> str:
    report_date = _parse_report_date(base_date_str)
    if report_date is not None:
        return f"QF \ud53d\uc5c5 {report_date.strftime('%Y-%m-%d')} \ub9c8\uac10\uc790\ub8cc(\ub098\uc2e4\ud328\ubc00\ub9ac).xlsx"
    return "QF \ud53d\uc5c5 \ub9c8\uac10\uc790\ub8cc(\ub098\uc2e4\ud328\ubc00\ub9ac).xlsx"


__all__ = [
    "CELL_MAP",
    "REPORT_TEMPLATE_SHEET_NAME",
    "ROW_LAYOUT",
    "TEMPLATE_REPORT_XLSX",
    "build_report_export_df",
    "build_report_export_filename",
    "build_report_export_payload",
    "fill_report_template",
    "make_report_xlsx_bytes",
]
